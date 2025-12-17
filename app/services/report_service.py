"""Report Service - Handle report generation and exports"""

import uuid
import io
import csv
import json
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.report import Report, ReportTemplate, ReportSchedule, ReportStatusEnum, ExportFormatEnum
from app.models.user import User
from app.models.opportunity import Opportunity
from app.models.match import Match, Swipe
from app.models.payment import Transaction, PaymentProposal
from app.models.message import Message, Conversation
from app.models.learning import Course, Enrollment
from app.models.analytics import UserActivity, PlatformMetrics


class ReportService:
    """Service layer for report generation and data export"""

    @staticmethod
    async def generate_report_data(
        db: AsyncSession,
        report_type: str,
        fields: List[str],
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_order: str = "asc",
        limit: Optional[int] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        """
        Generate report data based on type and parameters

        Returns:
            Tuple of (data_rows, total_count)
        """

        # Map report types to models and default fields
        model_map = {
            "users": User,
            "opportunities": Opportunity,
            "matches": Match,
            "payments": Transaction,
            "messages": Message,
            "learning": Course,
            "analytics": PlatformMetrics
        }

        if report_type not in model_map:
            raise ValueError(f"Invalid report type: {report_type}")

        model = model_map[report_type]

        # Build base query
        query = select(model)

        # Apply filters
        if filters:
            filter_conditions = []
            for field, value in filters.items():
                if hasattr(model, field):
                    column = getattr(model, field)

                    # Handle different filter types
                    if isinstance(value, dict):
                        # Range filters: {"gte": 100, "lte": 1000}
                        if "gte" in value:
                            filter_conditions.append(column >= value["gte"])
                        if "lte" in value:
                            filter_conditions.append(column <= value["lte"])
                        if "gt" in value:
                            filter_conditions.append(column > value["gt"])
                        if "lt" in value:
                            filter_conditions.append(column < value["lt"])
                        if "eq" in value:
                            filter_conditions.append(column == value["eq"])
                        if "in" in value:
                            filter_conditions.append(column.in_(value["in"]))
                    else:
                        # Direct equality
                        filter_conditions.append(column == value)

            if filter_conditions:
                query = query.where(and_(*filter_conditions))

        # Get total count
        count_query = select(func.count()).select_from(model)
        if filters:
            count_query = count_query.where(and_(*filter_conditions))

        count_result = await db.execute(count_query)
        total_count = count_result.scalar() or 0

        # Apply sorting
        if sort_by and hasattr(model, sort_by):
            column = getattr(model, sort_by)
            if sort_order.lower() == "desc":
                query = query.order_by(column.desc())
            else:
                query = query.order_by(column.asc())
        else:
            # Default sort by created_at if available
            if hasattr(model, "created_at"):
                query = query.order_by(model.created_at.desc())

        # Apply limit
        if limit:
            query = query.limit(limit)

        # Execute query
        result = await db.execute(query)
        records = result.scalars().all()

        # Convert to dictionaries with only requested fields
        data_rows = []
        for record in records:
            row = {}
            for field in fields:
                if hasattr(record, field):
                    value = getattr(record, field)
                    # Handle special types
                    if isinstance(value, datetime):
                        row[field] = value.isoformat()
                    elif isinstance(value, uuid.UUID):
                        row[field] = str(value)
                    elif isinstance(value, (dict, list)):
                        row[field] = value
                    else:
                        row[field] = value
            data_rows.append(row)

        return data_rows, total_count

    @staticmethod
    async def export_to_csv(data: List[Dict[str, Any]], fields: List[str]) -> bytes:
        """Export data to CSV format"""
        output = io.StringIO()

        if not data:
            return b""

        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()

        for row in data:
            # Ensure all fields are present
            csv_row = {field: row.get(field, "") for field in fields}
            writer.writerow(csv_row)

        return output.getvalue().encode("utf-8")

    @staticmethod
    async def export_to_json(data: List[Dict[str, Any]]) -> bytes:
        """Export data to JSON format"""
        json_str = json.dumps(data, indent=2, default=str)
        return json_str.encode("utf-8")

    @staticmethod
    async def export_to_excel(data: List[Dict[str, Any]], fields: List[str]) -> bytes:
        """
        Export data to Excel format
        Note: Requires openpyxl library
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Report Data"

            # Write header
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=1, column=col_idx, value=field)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, field in enumerate(fields, start=1):
                    value = row_data.get(field, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        except ImportError:
            raise ValueError("openpyxl library required for Excel export. Install with: pip install openpyxl")

    @staticmethod
    async def export_to_pdf(data: List[Dict[str, Any]], fields: List[str], report_name: str) -> bytes:
        """
        Export data to PDF format
        Note: Requires reportlab library
        """
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

            output = io.BytesIO()

            # Create PDF document
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()

            # Add title
            title = Paragraph(f"<b>{report_name}</b>", styles["Title"])
            elements.append(title)
            elements.append(Spacer(1, 12))

            # Add timestamp
            timestamp = Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                styles["Normal"]
            )
            elements.append(timestamp)
            elements.append(Spacer(1, 20))

            # Prepare table data
            table_data = [fields]  # Header row

            for row in data:
                table_row = [str(row.get(field, "")) for field in fields]
                table_data.append(table_row)

            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)

            # Build PDF
            doc.build(elements)
            output.seek(0)
            return output.read()

        except ImportError:
            raise ValueError("reportlab library required for PDF export. Install with: pip install reportlab")

    @staticmethod
    async def create_report_from_template(
        db: AsyncSession,
        template: ReportTemplate,
        user_id: uuid.UUID,
        export_format: ExportFormatEnum,
        parameter_overrides: Optional[Dict[str, Any]] = None
    ) -> Report:
        """Generate a report from a template"""

        # Merge template filters with parameter overrides
        filters = template.filters or {}
        if parameter_overrides:
            filters.update(parameter_overrides)

        # Create report record
        report = Report(
            id=uuid.uuid4(),
            name=f"{template.name} - {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
            report_type=template.report_type,
            template_id=template.id,
            generated_by_id=user_id,
            export_format=export_format,
            parameters={
                "fields": template.fields,
                "filters": filters,
                "sort_by": template.sort_by,
                "sort_order": template.sort_order,
            },
            status=ReportStatusEnum.pending,
            progress_percentage=0,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )

        db.add(report)
        await db.commit()
        await db.refresh(report)

        return report

    @staticmethod
    async def process_report(db: AsyncSession, report: Report) -> Report:
        """Process a report - generate data and export file"""

        try:
            # Update status to processing
            report.status = ReportStatusEnum.processing
            report.progress_percentage = 10
            await db.commit()

            # Extract parameters
            params = report.parameters or {}
            fields = params.get("fields", [])
            filters = params.get("filters")
            sort_by = params.get("sort_by")
            sort_order = params.get("sort_order", "asc")

            # Generate data
            report.progress_percentage = 30
            await db.commit()

            data, total_count = await ReportService.generate_report_data(
                db=db,
                report_type=report.report_type,
                fields=fields,
                filters=filters,
                sort_by=sort_by,
                sort_order=sort_order
            )

            report.total_records = total_count
            report.progress_percentage = 60
            await db.commit()

            # Export to requested format
            export_data = None
            file_extension = ""

            if report.export_format == ExportFormatEnum.csv:
                export_data = await ReportService.export_to_csv(data, fields)
                file_extension = "csv"
            elif report.export_format == ExportFormatEnum.json:
                export_data = await ReportService.export_to_json(data)
                file_extension = "json"
            elif report.export_format == ExportFormatEnum.excel:
                export_data = await ReportService.export_to_excel(data, fields)
                file_extension = "xlsx"
            elif report.export_format == ExportFormatEnum.pdf:
                export_data = await ReportService.export_to_pdf(data, fields, report.name)
                file_extension = "pdf"

            report.progress_percentage = 80
            await db.commit()

            # In production, save file to cloud storage (S3, etc.)
            # For now, we'll store the file path placeholder
            file_name = f"report_{report.id}.{file_extension}"
            file_path = f"/tmp/reports/{file_name}"

            # Calculate data snapshot (summary statistics)
            data_snapshot = {
                "total_records": total_count,
                "fields_included": fields,
                "filters_applied": filters or {},
                "export_size_bytes": len(export_data) if export_data else 0,
                "generated_at": datetime.utcnow().isoformat()
            }

            # Update report with results
            report.file_path = file_path
            report.file_size_bytes = len(export_data) if export_data else 0
            report.download_url = f"/api/v1/reports/{report.id}/download"
            report.expires_at = datetime.utcnow() + timedelta(days=7)  # 7-day expiry
            report.data_snapshot = data_snapshot
            report.generated_at = datetime.utcnow()
            report.status = ReportStatusEnum.completed
            report.progress_percentage = 100

            await db.commit()
            await db.refresh(report)

            return report

        except Exception as e:
            # Update report with error
            report.status = ReportStatusEnum.failed
            report.error_message = str(e)
            await db.commit()
            raise

    @staticmethod
    async def process_report_background(report_id: uuid.UUID) -> None:
        """Process a report in background - creates its own database session"""
        from app.core.database import get_session_maker

        session_maker = get_session_maker()
        async with session_maker() as db:
            try:
                # Fetch the report
                query = select(Report).where(Report.id == report_id)
                result = await db.execute(query)
                report = result.scalar_one_or_none()

                if not report:
                    return

                # Process the report
                await ReportService.process_report(db, report)

            except Exception as e:
                # Log the error (in production, use proper logging)
                print(f"Error processing report {report_id}: {e}")
                # Try to update report status to failed
                try:
                    query = select(Report).where(Report.id == report_id)
                    result = await db.execute(query)
                    report = result.scalar_one_or_none()
                    if report:
                        report.status = ReportStatusEnum.failed
                        report.error_message = str(e)
                        await db.commit()
                except:
                    pass

    @staticmethod
    async def calculate_next_run_time(
        schedule: ReportSchedule,
        from_time: Optional[datetime] = None
    ) -> datetime:
        """Calculate next run time for a scheduled report"""

        if from_time is None:
            from_time = datetime.utcnow()

        # Parse time_of_day (HH:MM format)
        hour, minute = map(int, schedule.time_of_day.split(":"))

        # Start from tomorrow at the scheduled time
        next_run = from_time.replace(hour=hour, minute=minute, second=0, microsecond=0)

        if next_run <= from_time:
            # If time has passed today, start from tomorrow
            next_run += timedelta(days=1)

        # Adjust based on frequency
        if schedule.frequency == "daily":
            # Already set correctly
            pass

        elif schedule.frequency == "weekly":
            # Move to next occurrence of day_of_week
            target_weekday = schedule.day_of_week or 0  # Monday default
            days_ahead = target_weekday - next_run.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            next_run += timedelta(days=days_ahead)

        elif schedule.frequency == "monthly":
            # Move to next occurrence of day_of_month
            target_day = schedule.day_of_month or 1
            next_run = next_run.replace(day=target_day)
            if next_run <= from_time:
                # Move to next month
                if next_run.month == 12:
                    next_run = next_run.replace(year=next_run.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=next_run.month + 1)

        elif schedule.frequency == "quarterly":
            # Move to next quarter
            current_quarter = (next_run.month - 1) // 3
            next_quarter_month = (current_quarter + 1) * 3 + 1
            if next_quarter_month > 12:
                next_run = next_run.replace(year=next_run.year + 1, month=1)
            else:
                next_run = next_run.replace(month=next_quarter_month)

        elif schedule.frequency == "yearly":
            # Move to next year
            next_run = next_run.replace(year=next_run.year + 1)

        return next_run
